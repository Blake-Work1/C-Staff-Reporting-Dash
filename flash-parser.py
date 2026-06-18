"""
flash_parser.py
───────────────
Parses the Bookings & Renewals Flash Excel file into a standardised
ACV DataFrame that can be used directly by the dashboard or stored
in the database.

Usage:
    from flash_parser import parse_flash_file
    df = parse_flash_file("Bookings_Renewals_Flash_May_2026_FINAL.xlsx")

Output columns (one row per opportunity):
    ucid, account_name, uoid, sfdc_opp_id,
    eom, close_date,
    geo, tier, segment, fulfillment_channel, new_vs_existing, deal_type,
    nacv_usd, nacv_uplift_usd, total_acv_usd,
    tosca_bi_sl, tosca_bi_cpi,
    tosca_sl, tosca_cpi,
    tosca_osv_sl, tosca_osv_cpi,
    tee_sl, tee_cpi,
    tta_sl, tta_cpi,
    testim_sf_sl, testim_sf_cpi,
    neoload_sl, neoload_cpi,
    qtest_sl, qtest_cpi,
    livecompare_sl, livecompare_cpi,
    testim_sl, testim_cpi,
    vera_sl, vera_cpi,
    mobile_sl, mobile_cpi,
    tdc_sl, tdc_cpi,
    sealights_sl, sealights_cpi,
    agentic_sl, agentic_cpi,
    advisory_services_acv,
    support_acv,
    stage,
    is_closed_won,
    is_closed_pending,
    data_source
"""

from __future__ import annotations
import re
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# ── CONSTANTS ────────────────────────────────────────────────────────────────

CLOSED_WON_STAGES     = {"Closed Won", "Stage 5 - Closed Won"}
CLOSED_PENDING_STAGES = {"Stage 4 - Closed Pending", "6 - Closed/Pending"}
ALL_CLOSED_STAGES     = CLOSED_WON_STAGES | CLOSED_PENDING_STAGES

# Maps human-readable product name → (sales_led_col_header, cpi_col_header)
PRODUCT_COL_MAP = {
    "tosca_bi":    ("Tosca BI Sales Led",              "Tosca BI CPI"),
    "tosca":       ("ToscaSales Led",                  "Tosca CPI"),
    "tosca_osv":   ("Tosca OSV Sales Led",             "Tosca OSV CPI"),
    "tee":         ("TEE Sales Led",                   "TEE CPI"),
    "tta":         ("TTA Sales Led",                   "TTA CPI"),
    "testim_sf":   ("Testim Salesforce Sales Led",     "Testim Salesforce CPI"),
    "neoload":     ("NeoLoad Sales Led",               "NeoLoad CPI"),
    "qtest":       ("qTest Sales Led",                 "qTest CPI"),
    "livecompare": ("LiveCompare Sales Led",           "LiveCompare CPI"),
    "testim":      ("Testim Sales Led",                "Testim CPI"),
    "vera":        ("Vera Sales Led",                  "VERA CPI"),
    "mobile":      ("Mobile Sales Led",                "Mobile CPI"),
    "tdc":         ("TDC Sales Led",                   "TDC CPI"),
    "sealights":   ("SeaLights Sales Led",             "SeaLights CPI"),
    "agentic":     ("Agentic Sales Led",               "Agentic CPI"),
}

# Single-column products
SINGLE_COL_MAP = {
    "advisory_services_acv": "Advisory Services Acv",
    "support_acv":           "Support ACV",
}

# Core identifier / classification columns
IDENTITY_COLS = {
    "ucid":               "UCID",
    "account_name":       "Account Name",
    "uoid":               "UOID",
    "sfdc_opp_id":        "SFDC Opportunity ID (18)",
    "tier":               "Current Segment (Tier)",
    "enterprise_flag":    "Enterprise vs. Non",
    "geo":                "GEO_ADJ",
    "segment":            "Segment",
    "fulfillment_channel":"Fulfillment Channel",
    "new_vs_existing":    "New vs. Existing",
    "deal_type":          "Type",
    "stage":              "Stage",
    "close_date":         "Close Date",
    "eom":                "EOM",
    "nacv_usd":           "NACV (converted)",
    "nacv_uplift_usd":    "NACV Uplift (converted)",
    "total_acv_usd":      "Total ACV",
}


# ── HELPERS ──────────────────────────────────────────────────────────────────

def _to_float(val) -> float:
    """Safely convert a cell value to float; return 0.0 on failure."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _to_date(val) -> date | None:
    """Convert Excel date value to Python date."""
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(val[:10], fmt[:len(val[:10])]).date()
            except ValueError:
                pass
    return None


def _build_col_index(header_row) -> dict[str, int]:
    """Build a column-name → index map from the header row."""
    return {
        str(h).strip(): i
        for i, h in enumerate(header_row)
        if h is not None
    }


# ── MAIN PARSER ──────────────────────────────────────────────────────────────

def parse_flash_file(
    filepath: str | Path,
    sheet: str = "2026 Data",
    cutoff_date: date | None = None,
) -> pd.DataFrame:
    """
    Parse the Flash Excel file and return a standardised ACV DataFrame.

    Parameters
    ----------
    filepath     : Path to the Flash .xlsx file
    sheet        : Sheet name containing raw deal rows (default: "2026 Data")
    cutoff_date  : Only include EOM periods <= this date.
                   Defaults to today — future months are always excluded.

    Returns
    -------
    pd.DataFrame with one row per opportunity, closed-won and pending flags,
    and all product-level ACV columns.
    """
    if cutoff_date is None:
        cutoff_date = date.today()

    wb = load_workbook(filepath, read_only=True, data_only=True)

    if sheet not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet]
    all_rows = list(ws.iter_rows(values_only=True))

    # Row 0 = top-level group labels (partial)
    # Row 1 = actual column headers
    header_row = all_rows[1]
    col = _build_col_index(header_row)

    # ── Validate all required columns exist ──
    missing = []
    for friendly, raw in IDENTITY_COLS.items():
        if raw not in col:
            missing.append(raw)
    for prod, (sl, cpi) in PRODUCT_COL_MAP.items():
        if sl not in col:
            missing.append(sl)
        if cpi not in col:
            missing.append(cpi)
    for friendly, raw in SINGLE_COL_MAP.items():
        if raw not in col:
            missing.append(raw)

    if missing:
        raise ValueError(
            f"The following expected columns were not found in '{sheet}':\n"
            + "\n".join(f"  - {m}" for m in missing)
        )

    # ── Parse rows ──
    records = []

    for r in all_rows[2:]:  # skip both header rows
        # Skip completely empty rows
        if not any(r):
            continue

        stage_raw = str(r[col["Stage"]]).strip() if r[col["Stage"]] else ""

        # Only process closed stages
        if stage_raw not in ALL_CLOSED_STAGES:
            continue

        # Check EOM — exclude future periods
        eom = _to_date(r[col["EOM"]])
        if eom is None or eom > cutoff_date:
            continue

        # ── Identity / classification ──
        record = {
            "ucid":               str(r[col["UCID"]]).strip() if r[col["UCID"]] else None,
            "account_name":       str(r[col["Account Name"]]).strip() if r[col["Account Name"]] else None,
            "uoid":               str(r[col["UOID"]]).strip() if r[col["UOID"]] else None,
            "sfdc_opp_id":        str(r[col["SFDC Opportunity ID (18)"]]).strip() if r[col["SFDC Opportunity ID (18)"]] else None,
            "eom":                eom,
            "close_date":         _to_date(r[col["Close Date"]]),
            "geo":                str(r[col["GEO_ADJ"]]).strip() if r[col["GEO_ADJ"]] else None,
            "tier":               str(r[col["Current Segment (Tier)"]]).strip() if r[col["Current Segment (Tier)"]] else None,
            "enterprise_flag":    str(r[col["Enterprise vs. Non"]]).strip() if r[col["Enterprise vs. Non"]] else None,
            "segment":            str(r[col["Segment"]]).strip() if r[col["Segment"]] else None,
            "fulfillment_channel":str(r[col["Fulfillment Channel"]]).strip() if r[col["Fulfillment Channel"]] else None,
            "new_vs_existing":    str(r[col["New vs. Existing"]]).strip() if r[col["New vs. Existing"]] else None,
            "deal_type":          str(r[col["Type"]]).strip() if r[col["Type"]] else None,
            "stage":              stage_raw,
            "is_closed_won":      stage_raw in CLOSED_WON_STAGES,
            "is_closed_pending":  stage_raw in CLOSED_PENDING_STAGES,
        }

        # ── Financial ──
        record["nacv_usd"]        = _to_float(r[col["NACV (converted)"]])
        record["nacv_uplift_usd"] = _to_float(r[col["NACV Uplift (converted)"]])
        record["total_acv_usd"]   = _to_float(r[col["Total ACV"]])

        # ── Product columns ──
        for prod_key, (sl_hdr, cpi_hdr) in PRODUCT_COL_MAP.items():
            record[f"{prod_key}_sl"]  = _to_float(r[col[sl_hdr]])
            record[f"{prod_key}_cpi"] = _to_float(r[col[cpi_hdr]])

        for field_key, hdr in SINGLE_COL_MAP.items():
            record[field_key] = _to_float(r[col[hdr]])

        # ── Source tag ──
        record["data_source"] = "flash"

        records.append(record)

    df = pd.DataFrame(records)

    if df.empty:
        print("WARNING: No rows matched the filter criteria.")
        return df

    # ── Post-processing ──

    # Clean up fulfillment channel — treat #N/A as Unknown
    df["fulfillment_channel"] = df["fulfillment_channel"].replace("#N/A", "Unknown")

    # Standardise tier labels
    tier_map = {
        "Tier 1": "Tier 1",
        "Tier 2": "Tier 2",
        "Tier 3": "Tier 3",
        "Enterprise":     "Tier 1/2",   # fallback if tier col is blank
        "Non-Enterprise": "Tier 3",
    }
    df["tier"] = df["tier"].map(lambda t: tier_map.get(t, t) if t else None)

    # Ensure eom is datetime for easy filtering downstream
    df["eom"] = pd.to_datetime(df["eom"])

    print(f"Flash parser complete:")
    print(f"  File     : {Path(filepath).name}")
    print(f"  Sheet    : {sheet}")
    print(f"  Cutoff   : {cutoff_date}")
    print(f"  Rows in  : {len(all_rows) - 2}")
    print(f"  Rows out : {len(df)}")
    print(f"  Periods  : {sorted(df['eom'].dt.strftime('%Y-%m-%d').unique())}")
    print(f"  CW rows  : {df['is_closed_won'].sum()}")
    print(f"  CP rows  : {df['is_closed_pending'].sum()}")

    return df


# ── AGGREGATION HELPERS ───────────────────────────────────────────────────────

def acv_by_geo(df: pd.DataFrame, closed_won_only: bool = True) -> pd.DataFrame:
    """
    Aggregate total ACV by geo and period.
    Returns columns: eom, geo, total_acv_usd, nacv_usd
    """
    mask = df["is_closed_won"] if closed_won_only else slice(None)
    grp = df[mask].groupby(
        [df["eom"].dt.to_period("M"), "geo"],
        as_index=False
    ).agg(
        total_acv_usd=("total_acv_usd", "sum"),
        nacv_usd=("nacv_usd", "sum"),
        deal_count=("uoid", "count"),
    )
    grp.rename(columns={"eom": "period"}, inplace=True)
    return grp.sort_values(["period", "geo"])


def acv_by_product(df: pd.DataFrame, closed_won_only: bool = True) -> pd.DataFrame:
    """
    Aggregate Sales Led and CPI ACV by product and period.
    Returns a long-format DataFrame: period, product, sales_led, cpi, total
    """
    mask = df["is_closed_won"] if closed_won_only else slice(None)
    sub = df[mask].copy()
    sub["period"] = sub["eom"].dt.to_period("M")

    products = list(PRODUCT_COL_MAP.keys())
    records = []
    for prod in products:
        agg = sub.groupby("period").agg(
            sales_led=(f"{prod}_sl", "sum"),
            cpi=(f"{prod}_cpi", "sum"),
        ).reset_index()
        agg["product"] = prod
        agg["total"] = agg["sales_led"] + agg["cpi"]
        records.append(agg)

    result = pd.concat(records, ignore_index=True)
    return result[["period", "product", "sales_led", "cpi", "total"]].sort_values(
        ["period", "product"]
    )


def acv_by_new_existing(df: pd.DataFrame, closed_won_only: bool = True) -> pd.DataFrame:
    """
    Aggregate ACV split by New vs Existing customer and geo.
    """
    mask = df["is_closed_won"] if closed_won_only else slice(None)
    grp = df[mask].groupby(
        [df["eom"].dt.to_period("M"), "geo", "new_vs_existing"],
        as_index=False
    ).agg(total_acv_usd=("total_acv_usd", "sum"))
    grp.rename(columns={"eom": "period"}, inplace=True)
    return grp.sort_values(["period", "geo", "new_vs_existing"])


# ── QUICK TEST ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else \
        "/mnt/user-data/uploads/Bookings___Renewals_Flash_-_May_2026_FINAL.xlsx"

    df = parse_flash_file(path)

    print()
    print("=== ACV by Geo (Closed Won only) ===")
    print(acv_by_geo(df).to_string(index=False))

    print()
    print("=== ACV by Product (Closed Won, summed across all periods) ===")
    prod = acv_by_product(df)
    totals = prod.groupby("product")[["sales_led","cpi","total"]].sum()
    totals = totals[totals["total"].abs() > 0].sort_values("total", ascending=False)
    print(totals.to_string())

    print()
    print("=== New vs Existing by Geo ===")
    print(acv_by_new_existing(df).to_string(index=False))
